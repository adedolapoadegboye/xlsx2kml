#! /usr/bin/env python
# -*- python -*-

from Tkinter import *
import ttk

from tkFileDialog import askopenfilename, askdirectory
from tkMessageBox import *

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import coordinate_from_string, \
    column_index_from_string, get_column_letter
    
from pysqlite2 import dbapi2 as db 

import webbrowser

def vp_start_gui():
    
    global val, w, root
    root = Tk()
    root.title('XY2KML')
    root.geometry('402x435+259+154')
    root.iconbitmap('logo_xy2kml.ico')
    root.resizable(0,0)
    set_Tk_var()
    w = New_Toplevel_1 (root)
    init()
    root.mainloop()

w = None
def create_New_Toplevel_1 (root):
    
    global w, w_win
    if w: 
        return
    w = Toplevel (root)
    w.title('XY2KML')
    w.geometry('402x435+259+154')
    w.iconbitmap('logo_xy2kml.ico')
    set_Tk_var()
    w_win = New_Toplevel_1 (w)
    init()
    return w_win

def destroy_New_Toplevel_1 ():
    global w
    w.destroy()
    w = None


def set_Tk_var():
    
    global combobox, combobox2, combobox3, combobox4, combobox5
    combobox = StringVar()
    combobox2 = StringVar()
    combobox3 = StringVar()
    combobox4 = StringVar()
    combobox5 = StringVar()

    global tch33
    tch33 = IntVar()


def init():
    pass
    
class popupAbout(object):

    def about_link_callback(event, *args):
        webbrowser.open_new("http://yilmazturk.info/")
        
    def github_link_callback(event, *args):
        webbrowser.open_new("https://github.com/syilmazturk/xy2kml")

    def __init__(self, master=None):
        
        top=self.top=Toplevel(master)
        top.geometry('250x207+598+259')
        top.iconbitmap('logo_xy2kml.ico')
        top.title('XY2KML | About')
        top.resizable(0,0)
        
        self.TLabel1 = ttk.Label (top)
        self.TLabel1.place(relx=0.43,rely=0.19,height=19,width=72)
        self.TLabel1.configure(relief="flat")
        self.TLabel1.configure(text='''XY2KML v1.0''')

        self.TLabel2 = ttk.Label (top)
        self.TLabel2.place(relx=0.28,rely=0.43,height=19,width=113)
        self.TLabel2.configure(relief="flat")
        
        self.TLabel2.configure(text='''www.yilmazturk.info''', \
        foreground="blue", cursor="hand2")
        
        self.TLabel2.bind("<Button-1>", self.about_link_callback)

        self.TLabel3 = ttk.Label (top)
        self.TLabel3.place(relx=0.16,rely=0.58,height=19,width=181)
        self.TLabel3.configure(relief="flat")
        
        self.TLabel3.configure(text='''github.com/syilmazturk/xy2kml''', \
        foreground="blue", cursor="hand2")
        
        self.TLabel3.bind("<Button-1>", self.github_link_callback)

        self.TButton1 = ttk.Button (top)
        self.TButton1.place(relx=0.36,rely=0.77,height=25,width=76)
        self.TButton1.configure(takefocus="")
        self.TButton1.configure(text='''OK''', command = top.destroy)

        self.logoXY2KML = PhotoImage(file='logo_xy2kml_36.gif')
        
        self.TLabel2 = ttk.Label (top)
        self.TLabel2.place(relx=0.24,rely=0.14,height=40,width=36)
        self.TLabel2.configure(relief="flat")
        self.TLabel2.configure(text='''''', image=self.logoXY2KML)
        
        
class popupWindow(object):
    
    #specify crs for excel document that contains coordinate values
    
    def show_ref_sys_name(self, *args):
    
        conn = db.connect('db.sqlite')
        conn.text_factory = str
        
        cursor = conn.execute("SELECT ref_sys_name FROM spatial_ref_sys \
        WHERE srid = %s " % self.TCombobox5.get()) 
        
        results = cursor.fetchall()
        self.TLabel2.configure(text='''%s''' % str(results)[3:-4])
           
    
    def __init__(self, master=None):
        
        top=self.top=Toplevel(master)
        top.geometry('351x123+458+222')
        top.iconbitmap('logo_xy2kml.ico')
        top.title('XY2KML | Specify CRS')
        top.resizable(0,0)

        self.TCombobox5 = ttk.Combobox (top, state='readonly')
        self.TCombobox5.place(relx=0.25,rely=0.16,relheight=0.17,relwidth=0.49)
        self.TCombobox5.configure(textvariable=combobox5)
        self.TCombobox5.configure(takefocus="")
        self.TCombobox5.set("Specify Input CRS...")
        self.TCombobox5.bind("<<ComboboxSelected>>", self.show_ref_sys_name)

        self.TButton1 = ttk.Button (top, command = self.get_epsg_value)
        self.TButton1.place(relx=0.39,rely=0.41,height=25,width=76)
        self.TButton1.configure(takefocus="")
        self.TButton1.configure(text='''OK''')

        self.TLabel2 = ttk.Label (top, anchor=CENTER)
        self.TLabel2.place(relx=0.03,rely=0.73,height=19,width=339)
        self.TLabel2.configure(relief="flat")
        self.TLabel2.configure(text='''-''')
        
        conn = db.connect('db.sqlite')
        conn.enable_load_extension(True)
        conn.load_extension("libspatialite-4.dll")
        
        cursor = conn.execute("SELECT srid FROM spatial_ref_sys \
        ORDER BY srid ASC;")
        
        results = cursor.fetchall()
        crs = [foo[0] for foo in results]
        self.TCombobox5['values'] = crs
        
    def get_epsg_value(self):
        
        self.epsg_value = self.TCombobox5.get()
        self.top.destroy()


class New_Toplevel_1(object):

    def open_excel_file(self):
        
        global workbook
        
        excel_file = askopenfilename(title = "Select Excel File...", \
        filetypes = [('Excel Document', '*.xlsx')])
        
        if excel_file == "":
            showwarning('Fail...', '.XLSX is not selected!')    
        else:    
            excel_file_name = excel_file.split('/')[-1]
            showinfo('Success...', excel_file_name + ' is selected.')
            workbook = load_workbook(excel_file)
            sheet_names = [foo.title for foo in workbook]
            self.TCombobox1['values'] = sheet_names
            
    def get_xy_columns(self, *args):
    
    #returns excel document's first row values as column names
    #get_highest_column method yields total column number which contains value
    #keys list contains columns as letters like A, B, C etc.
    #values list contains first row values as cell-by-cell    
        
        global get_sheet, keys, values
        
        get_sheet = workbook.get_sheet_by_name(self.TCombobox1.get())
        
        ghc = get_sheet.get_highest_column() + 1

        keys = []
        values = [] 
        
        for col_idx in range(1, ghc):
            col = get_column_letter(col_idx)
            keys.append(col)
            values.append(get_sheet.cell(col + '1').value)
        
        self.TCombobox2['values'] = values
        self.TCombobox3['values'] = values

    def __init__(self, master=None):
        
        self.master=master
        style = ttk.Style()
        theme = style.theme_use()
        default = style.lookup(theme, 'background')
        master.configure(background=default)

        self.TLabelframe1 = ttk.Labelframe (master)
        self.TLabelframe1.place(relx=0.02,rely=0.07,relheight=0.63
                ,relwidth=0.96)
        self.TLabelframe1.configure(text='''Specify the parameters for data to be converted to KML''')

        self.TLabel1 = ttk.Label (self.TLabelframe1)
        self.TLabel1.place(relx=0.03,rely=0.58,height=19,width=128)
        self.TLabel1.configure(relief="flat")
        self.TLabel1.configure(text='''X / Easting / Longitude:''')

        self.TCombobox1 = ttk.Combobox (self.TLabelframe1, state='readonly')
        self.TCombobox1.place(relx=0.42,rely=0.25,relheight=0.08,relwidth=0.55)
        self.TCombobox1.configure(textvariable=combobox)
        self.TCombobox1.configure(takefocus="")
        self.TCombobox1.set("Select Sheet...")
        self.TCombobox1.bind("<<ComboboxSelected>>", self.get_xy_columns)

        self.TLabel2 = ttk.Label (self.TLabelframe1)
        self.TLabel2.place(relx=0.03,rely=0.73,height=19,width=127)
        self.TLabel2.configure(relief="flat")
        self.TLabel2.configure(text='''Y / Northing / Latitude:''')

        self.TCombobox2 = ttk.Combobox (self.TLabelframe1, state='readonly')
        self.TCombobox2.place(relx=0.42,rely=0.58,relheight=0.08,relwidth=0.55)
        self.TCombobox2.configure(textvariable=combobox2)
        self.TCombobox2.configure(takefocus="")
        self.TCombobox2.set("Select X Column...")

        self.TCombobox3 = ttk.Combobox (self.TLabelframe1, state='readonly')
        self.TCombobox3.place(relx=0.42,rely=0.73,relheight=0.08,relwidth=0.55)
        self.TCombobox3.configure(textvariable=combobox3)
        self.TCombobox3.configure(takefocus="")
        self.TCombobox3.set("Select Y Column...")

        self.TLabel3 = ttk.Label (self.TLabelframe1)
        self.TLabel3.place(relx=0.03,rely=0.44,height=19,width=253)
        self.TLabel3.configure(relief="flat")
        self.TLabel3.configure(text='''Select the columns for the X and Y coordinates:''')

        self.TLabel4 = ttk.Label (self.TLabelframe1)
        self.TLabel4.place(relx=0.03,rely=0.11,height=19,width=307)
        self.TLabel4.configure(relief="flat")
        self.TLabel4.configure(text='''Select the Excel sheet containing X and Y coordinate data:''')

        self.TButton1 = ttk.Button (master, command = self.run)
        self.TButton1.place(relx=0.75,rely=0.76,height=55,width=84)
        self.TButton1.configure(takefocus="")
        self.logoGE = PhotoImage(file='logo_ge.gif')
        self.TButton1.configure(text='''Generate KML''', \
        image=self.logoGE, compound="top")

        self.TButton2 = ttk.Button (master, command = self.popup)
        self.TButton2.place(relx=0.77,rely=0.62,height=25,width=76)
        self.TButton2.configure(takefocus="")
        self.TButton2.configure(text='''Set CRS''')

        self.TCheckbutton1 = ttk.Checkbutton (master, command = self.crs)
        self.TCheckbutton1.place(relx=0.05,rely=0.76,relheight=0.05
                ,relwidth=0.42)
        self.TCheckbutton1.configure(variable=tch33)
        self.TCheckbutton1.configure(takefocus="")
        self.TCheckbutton1.configure(text='''Transform & Write to Excel?''')

        self.TCombobox4 = ttk.Combobox (master, state='readonly')
        self.TCombobox4.place(relx=0.05,rely=0.83,relheight=0.05,relwidth=0.53)
        self.TCombobox4.configure(textvariable=combobox4)
        self.TCombobox4.configure(takefocus="")
        self.TCombobox4.set("Specify Output CRS...")
        self.TCombobox4.bind("<<ComboboxSelected>>", self.show_ref_sys_name)

        self.TLabel5 = ttk.Label (master)
        self.TLabel5.place(relx=0.05,rely=0.92,height=19,width=31)
        self.TLabel5.configure(relief="flat")
        self.TLabel5.configure(text='''CRS: ''')

        self.TLabel6 = ttk.Label (master)
        self.TLabel6.place(relx=0.12,rely=0.92,height=19,width=349)
        self.TLabel6.configure(relief="flat")
        self.TLabel6.configure(text='''-''')

        self.build_menu()
        
    def build_menu(self):

        menubar = Menu(root)
        root.config(menu = menubar)
        filemenu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=filemenu)
        
        filemenu.add_command(label = 'Open...', command=self.open_excel_file)
        filemenu.add_separator()
        filemenu.add_command(label = 'Exit', command=root.destroy)
      
        helpmenu = Menu(menubar, tearoff=0)
        
        helpmenu.add_command(label="XY2KML Help", command=self.\
        help_link_callback)
        
        helpmenu.add_separator()
        helpmenu.add_command(label="About XY2KML...", command=self.popup_about)
        menubar.add_cascade(label="Help", menu=helpmenu)
        
    def help_link_callback(event, *args):
        webbrowser.open_new("http://yilmazturk.info/xy2kml/")    
        
    def generate_kml(self):
        
    #data_range_x & y specify the range that contains values like A2:A6    
        
        global conn, data_x, data_y, dir_name    
        
        first_row = dict(zip(values, keys))
        
        first_cell_x = first_row[self.TCombobox2.get()] + '2'
        last_cell_x = first_row[self.TCombobox2.get()] + \
        str(get_sheet.get_highest_row())
        data_range_x = first_cell_x + ":" + last_cell_x
        
        first_cell_y = first_row[self.TCombobox3.get()] + '2'
        
        last_cell_y = first_row[self.TCombobox3.get()] + \
        str(get_sheet.get_highest_row())
        
        data_range_y = first_cell_y + ":" + last_cell_y

        data_x = []
        data_y = []
        
        for row in get_sheet.range(data_range_x):
            for cell in row:
                data_x.append(cell.value)
        
        for row in get_sheet.range(data_range_y):
            for cell in row:
                data_y.append(cell.value)
            
        list_for_query = []    
            
        for i in range(0, get_sheet.get_highest_column() - 1):
            combine_x_y = '%s %s' % (data_x[i], data_y[i])
            list_for_query.append(combine_x_y)
            
        coord_list = ','.join(list_for_query)
            
        conn = db.connect('db.sqlite')
        conn.enable_load_extension(True)
        conn.load_extension("libspatialite-4.dll")
        conn.text_factory = str
        
        query_kml_generate = "SELECT AsKML('Point', 'Generated by XY2KML', \
        ST_GeomFromText('MULTIPOINT(%s)', %s))" % \
        (coord_list, self.epsg_popup.epsg_value)
        
        cursor = conn.execute(query_kml_generate)
        results = cursor.fetchall()
           
        dir_name = askdirectory()
            
        kml_file = open(dir_name + '/xy2kml.kml', 'w')
        kml_file.write(str(results)[3:-4])
        kml_file.close()
      
           
    def transform_xlsx(self):
    
        wb = Workbook()
        dest_filename = dir_name + '/transformed.xlsx'
        ws = wb.active
        ws.title = transformed_crs.replace('/', '-')

        query_result_set_x = []
        query_result_set_y = []
        
        for i in range(0, len(data_x)):
    
            query_generate_x = "SELECT ST_X(ST_Transform(ST_GeomFromText \
            ('POINT(%s %s)', %s), %s))" % (data_x[i], data_y[i], \
            self.epsg_popup.epsg_value, self.TCombobox4.get())
            
            cursor = conn.execute(query_generate_x)
            results_x = cursor.fetchall()
            query_result_set_x.append(str(results_x)[2:-3])
    
        for i in range(0, len(data_y)):
    
            query_generate_y = "SELECT ST_Y(ST_Transform(ST_GeomFromText \
            ('POINT(%s %s)', %s), %s))" % (data_x[i], data_y[i], \
            self.epsg_popup.epsg_value, self.TCombobox4.get())
            
            cursor = conn.execute(query_generate_y)
            results_y = cursor.fetchall()
            query_result_set_y.append(str(results_y)[2:-3])

        for col_idx in range(1, 2):
            col = get_column_letter(col_idx)
            for row in range(0, len(query_result_set_x)):
                ws.cell('%s%s'%(col, row + 1)).value = '%s' % \
                (query_result_set_x[row])   

        for col_idx in range(2, 3):
            col = get_column_letter(col_idx)
            for row in range(0, len(query_result_set_y)):
                ws.cell('%s%s'%(col, row + 1)).value = '%s' % \
                (query_result_set_y[row])           

        wb.save(filename = dest_filename)
        
        
    def crs(self):
    
        if tch33.get() == 1:
        
            conn = db.connect('db.sqlite')
            conn.enable_load_extension(True)
            conn.load_extension("libspatialite-4.dll")
            cursor = conn.execute \
            ("SELECT srid FROM spatial_ref_sys ORDER BY srid ASC;")
            results = cursor.fetchall()
            crs = [foo[0] for foo in results]
            self.TCombobox4['values'] = crs
            
        else:
        
            self.TCombobox4.set("Specify Output CRS...")
            self.TCombobox4['values'] = []
            self.TLabel6.configure(text='''-''')
            
            
    def show_ref_sys_name(self, *args):
    
        global transformed_crs
        
        conn = db.connect('db.sqlite')
        conn.text_factory = str
        
        cursor = conn.execute("SELECT ref_sys_name FROM spatial_ref_sys \
        WHERE srid = %s " % self.TCombobox4.get()) 
        
        results = cursor.fetchall()
        transformed_crs = str(results)[3:-4]
        self.TLabel6.configure(text='''%s''' % transformed_crs)
        
        
    def run(self):
    
        if tch33.get() == 1:
       
            self.generate_kml()
            self.transform_xlsx()
            
        else:
        
            self.generate_kml()
         
              
    def popup(self):
    
        self.epsg_popup = popupWindow(self.master)
        self.master.wait_window(self.epsg_popup.top)
        
    def popup_about(self):
    
        self.about_popup = popupAbout(self.master)
        self.master.wait_window(self.about_popup.top)


if __name__ == '__main__':
    vp_start_gui()



